import openpyxl;
import os;
from openpyxl.styles.borders import Border , Side;
wb = openpyxl.Workbook();
ws = wb.active;

ws.append(['Nume','Pret','Valuta','Cantitate','Unitate de masura','Suma'])

with open(rf"D:\Visual Studio\Projects\WidgetApp\depozitdir\info.txt" , "r") as file:
    name, price, valut, quantity, unit = " ", " ", " ", " ", " ";
    dir_ = " "
    filename_ = " "
    linecount = 0;
    lines = file.read().split("\n")
    for line in lines:
        if "=filename->" in line:
            filename_ = line.replace("=filename->", "")
            filename_ = filename_ + ".xlsx"
            continue;
        if "=dir->" in line:
            dir_ = line.replace("=dir->","")
            continue;
        if(line == ""):
            continue;
        linecount =linecount+1;
        name , price ,valut, quantity,unit = line.split("  " , 5);


        if(name != " "):
            ws.append([f'{(name)}', int(price) , f"{(valut)}.", int(quantity) ,f"{(unit)}."])
            ws[f"F{linecount + 1}"] = f"=PRODUCT(B{linecount + 1}:D{linecount + 1})";

    cell = ws.append([' ', ' ', ' '])
    ws.append(['Suma totala = ', ' '," "," "," ", f"=SUM(F2:F{linecount+1})"])
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'));
    for x in range(linecount+3): arg = f"A{x + 1}";ws[arg].border = thin_border;
    for x in range(linecount+3): arg = f"B{x + 1}";ws[arg].border = thin_border;
    for x in range(linecount+3): arg = f"C{x + 1}";ws[arg].border = thin_border;
    for x in range(linecount+3): arg = f"D{x + 1}";ws[arg].border = thin_border;
    for x in range(linecount + 3): arg = f"E{x + 1}";ws[arg].border = thin_border;
    for x in range(linecount + 3): arg = f"F{x + 1}";ws[arg].border = thin_border;

    file.close();

try:
    wb.save(rf'D:\Visual Studio\Projects\WidgetApp\projectdir\{dir_}\{filename_}');
except:
    absdir = rf"D:\Visual Studio\Projects\WidgetApp\projectdir\{dir_}";
    os.mkdir(absdir);
    wb.save(rf'D:\Visual Studio\Projects\WidgetApp\projectdir\{dir_}\{filename_}');


